import logging
from io import BytesIO

import openpyxl
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from courses.models import Session
from users.models import Classe, Student
from users.permissions import IsAdminOrTeacher, IsStudent
from users.roles import acting_as_admin, acting_as_teacher

from .models import Attendance, SuspiciousAttempt
from .serializers import (
    AttendanceSerializer,
    SuspiciousAttemptSerializer,
    ValidateAppAttendanceSerializer,
    ValidateAttendanceSerializer,
)
from .services import AttendanceService

logger = logging.getLogger(__name__)


def get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def get_user_agent(request):
    return request.META.get("HTTP_USER_AGENT", "")


def broadcast_attendance(attendance, session):
    try:
        channel_layer = get_channel_layer()
        payload = {
            "type": "attendance_update",
            "data": AttendanceSerializer(attendance).data,
        }
        async_to_sync(channel_layer.group_send)(f"attendance_{session.id}", payload)
        async_to_sync(channel_layer.group_send)("attendance_admin", payload)
        teacher_id = session.teacher_id
        if teacher_id:
            async_to_sync(channel_layer.group_send)(
                f"attendance_teacher_{teacher_id}", payload
            )
    except Exception as e:
        logger.warning(f"WebSocket broadcast failed: {e}")


class ValidateAttendanceThrottle(ScopedRateThrottle):
    scope = "validate_attendance"


class QrInfoRateThrottle(ScopedRateThrottle):
    """Dedicated per-IP quota for QR info checks, isolated from other
    anonymous traffic (e.g. login) sharing the same public IP."""

    scope = "qr_info"


class ValidateAttendanceView(APIView):
    permission_classes = [AllowAny]
    throttle_classes = [ValidateAttendanceThrottle]

    def post(self, request):
        serializer = ValidateAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        ip_address = get_client_ip(request)
        user_agent = get_user_agent(request)

        try:
            attendance, session = AttendanceService.validate_attendance(
                qr_token=data["qr_token"],
                first_name=data["first_name"],
                last_name=data["last_name"],
                code_massar=data["code_massar"],
                ip_address=ip_address,
                device_id=data.get("device_id", ""),
                device_fingerprint=data.get("device_fingerprint", ""),
                device_info=data.get("device_info", ""),
                latitude=data.get("latitude"),
                longitude=data.get("longitude"),
                user_agent=user_agent,
            )
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        broadcast_attendance(attendance, session)

        return Response(
            AttendanceSerializer(attendance).data, status=status.HTTP_201_CREATED
        )


class ValidateAppAttendanceView(APIView):
    """Authenticated mobile app validation — uses linked student profile."""

    permission_classes = [IsAuthenticated, IsStudent]
    throttle_classes = [ValidateAttendanceThrottle]

    def post(self, request):
        serializer = ValidateAppAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        ip_address = get_client_ip(request)
        user_agent = get_user_agent(request)

        try:
            attendance, session = AttendanceService.validate_attendance_for_user(
                qr_token=data["qr_token"],
                user=request.user,
                ip_address=ip_address,
                device_id=data.get("device_id", ""),
                device_fingerprint=data.get("device_fingerprint", ""),
                device_info=data.get("device_info", ""),
                latitude=data.get("latitude"),
                longitude=data.get("longitude"),
                user_agent=user_agent,
            )
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        broadcast_attendance(attendance, session)

        return Response(
            AttendanceSerializer(attendance).data, status=status.HTTP_201_CREATED
        )


class QrSessionInfoView(APIView):
    """Public endpoint to verify QR token validity before showing the web form."""

    permission_classes = [AllowAny]
    throttle_classes = [QrInfoRateThrottle]

    def get(self, request):
        token = request.query_params.get("token", "").strip()
        if not token:
            return Response(
                {"error": "Token requis."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            session = Session.objects.select_related("teacher", "classe").get(
                qr_token=token
            )
        except Session.DoesNotExist:
            return Response(
                {"valid": False, "error": "Code QR invalide."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "valid": session.is_qr_valid and session.is_active,
                "expired": not session.is_qr_valid,
                "session_id": session.id,
                "subject": session.subject,
                "classe": session.classe.name,
                "teacher": session.teacher.full_name,
                "date": session.date,
                "qr_expires_at": session.qr_expires_at,
                "qr_session_id": session.qr_session_id,
                "attendance_radius_meters": session.attendance_radius_meters,
                "requires_location": session.location_latitude is not None,
            }
        )


class MyAttendanceListView(ListAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, IsStudent]

    def get_queryset(self):
        user = self.request.user
        return Attendance.objects.select_related(
            "student__classe",
            "session__teacher",
            "session__classe",
        ).filter(student__user=user).order_by("-validation_time")


class AllAttendanceListView(ListAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get_queryset(self):
        qs = Attendance.objects.select_related(
            "student__classe",
            "session__teacher",
            "session__classe",
        ).order_by("-validation_time")

        user = self.request.user
        if acting_as_teacher(user):
            qs = qs.filter(session__teacher=user)

        session_id = self.request.query_params.get("session")
        classe_id = self.request.query_params.get("classe")
        if session_id:
            qs = qs.filter(session_id=session_id)
        if classe_id:
            qs = qs.filter(session__classe_id=classe_id)

        return qs


class SessionAttendanceListView(ListAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get_queryset(self):
        session_id = self.kwargs["session_id"]
        user = self.request.user

        if acting_as_teacher(user):
            try:
                session = Session.objects.select_related("teacher", "classe").get(
                    pk=session_id
                )
            except Session.DoesNotExist:
                return Attendance.objects.none()
            if not session.teacher_can_manage(user):
                return Attendance.objects.none()

        return Attendance.objects.select_related(
            "student__classe", "session__teacher", "session__classe"
        ).filter(session_id=session_id).order_by("-validation_time")


class SessionSuspiciousAttemptsView(ListAPIView):
    serializer_class = SuspiciousAttemptSerializer
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get_queryset(self):
        session_id = self.kwargs["session_id"]
        user = self.request.user

        try:
            session = Session.objects.get(pk=session_id)
        except Session.DoesNotExist:
            return SuspiciousAttempt.objects.none()

        if acting_as_teacher(user) and not session.teacher_can_manage(user):
            return SuspiciousAttempt.objects.none()

        return SuspiciousAttempt.objects.filter(session_id=session_id).order_by(
            "-created_at"
        )


class SessionStatsView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request, session_id):
        try:
            session = Session.objects.select_related("classe").get(pk=session_id)
        except Session.DoesNotExist:
            return Response(
                {"error": "Session not found."}, status=status.HTTP_404_NOT_FOUND
            )

        if not session.teacher_can_manage(request.user):
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        attendances = Attendance.objects.filter(session=session)
        suspicious = SuspiciousAttempt.objects.filter(session=session)
        total_students = session.classe.students.count()

        duplicate_attempts = suspicious.filter(
            attempt_type__in=[
                SuspiciousAttempt.ATTEMPT_DUPLICATE_STUDENT,
                SuspiciousAttempt.ATTEMPT_DUPLICATE_DEVICE,
            ]
        ).count()

        geofence_violations = suspicious.filter(
            attempt_type=SuspiciousAttempt.ATTEMPT_GEOFENCE
        ).count()

        devices_used = (
            attendances.exclude(device_fingerprint="")
            .values("device_fingerprint")
            .distinct()
            .count()
        )

        return Response(
            {
                "session_id": session.id,
                "subject": session.subject,
                "total_students_in_class": total_students,
                "attendance_count": attendances.count(),
                "attendance_rate": round(
                    (attendances.count() / total_students * 100)
                    if total_students
                    else 0,
                    1,
                ),
                "duplicate_attempts": duplicate_attempts,
                "geofence_violations": geofence_violations,
                "suspicious_attempts_total": suspicious.count(),
                "unique_devices": devices_used,
                "locations": [
                    {
                        "student_name": f"{a.student.first_name} {a.student.last_name}",
                        "code_massar": a.student.code_massar,
                        "latitude": a.latitude,
                        "longitude": a.longitude,
                        "validation_time": a.validation_time,
                    }
                    for a in attendances.filter(
                        latitude__isnull=False, longitude__isnull=False
                    )
                ],
            }
        )


class ExportAttendanceView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request, session_id):
        try:
            session = Session.objects.select_related("teacher", "classe").get(
                pk=session_id
            )
        except Session.DoesNotExist:
            return Response(
                {"error": "Session not found."}, status=status.HTTP_404_NOT_FOUND
            )

        if not session.teacher_can_manage(request.user):
            return Response({"error": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        attendances = Attendance.objects.select_related(
            "student__classe", "session"
        ).filter(session=session)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Présence - {session.subject[:20]}"

        headers = [
            "#",
            "Prénom",
            "Nom",
            "Code Massar",
            "Classe",
            "Matière",
            "Date séance",
            "Heure validation",
            "Adresse IP",
            "Empreinte appareil",
            "Latitude",
            "Longitude",
            "Statut",
            "QR Session ID",
        ]
        ws.append(headers)

        from openpyxl.styles import Alignment, Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for idx, att in enumerate(attendances, start=1):
            local_time = timezone.localtime(att.validation_time)
            ws.append(
                [
                    idx,
                    att.student.first_name,
                    att.student.last_name,
                    att.student.code_massar,
                    att.student.classe.name,
                    session.subject,
                    str(session.date),
                    local_time.strftime("%Y-%m-%d %H:%M:%S"),
                    att.ip_address or "",
                    att.device_fingerprint or att.device_id or "",
                    str(att.latitude) if att.latitude is not None else "",
                    str(att.longitude) if att.longitude is not None else "",
                    att.attendance_status,
                    att.qr_session_id,
                ]
            )

        for column in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in column), default=0)
            ws.column_dimensions[column[0].column_letter].width = max_len + 4

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_subject = "".join(c if c.isalnum() else "_" for c in session.subject)[:30]
        filename = f"presence_{safe_subject}_{session.date}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        logger.info(
            f"Exported attendance for session {session_id} by {request.user.email}"
        )
        return response


def _parse_id_list(raw: str) -> list[int]:
    if not raw or not raw.strip():
        return []
    return [int(part) for part in raw.split(",") if part.strip().isdigit()]


def _teacher_can_access_classe(user, classe) -> bool:
    if acting_as_admin(user):
        return True
    if not acting_as_teacher(user):
        return False
    return (
        Session.objects.filter(classe=classe)
        .filter(Q(teacher=user) | Q(module__teacher=user))
        .exists()
        or classe.semesters.filter(modules__teacher=user).exists()
    )


def _teacher_sessions_for_classe(user, classe, session_ids=None):
    qs = (
        Session.objects.select_related("module", "classe", "teacher")
        .filter(classe=classe)
        .order_by("-date", "start_time")
    )
    if acting_as_teacher(user) and not acting_as_admin(user):
        qs = qs.filter(Q(teacher=user) | Q(module__teacher=user))
    if session_ids:
        qs = qs.filter(pk__in=session_ids)
    return list(qs)


def _build_custom_export_data(user, classe_id, session_ids=None, student_ids=None):
    try:
        classe = Classe.objects.get(pk=classe_id)
    except Classe.DoesNotExist:
        raise ValueError("Classe introuvable.")

    if not _teacher_can_access_classe(user, classe):
        raise PermissionError("Accès refusé à cette classe.")

    sessions = _teacher_sessions_for_classe(user, classe, session_ids or None)
    if not sessions:
        raise ValueError("Aucune séance trouvée pour cette sélection.")

    students_qs = Student.objects.filter(classe=classe).order_by("last_name", "first_name")
    if student_ids:
        students_qs = students_qs.filter(pk__in=student_ids)
    students = list(students_qs)
    if not students:
        raise ValueError("Aucun étudiant trouvé pour cette sélection.")

    session_pks = [s.pk for s in sessions]
    attendances = Attendance.objects.select_related("student", "session").filter(
        session_id__in=session_pks,
        student_id__in=[s.pk for s in students],
        attendance_status="confirmed",
    )
    att_map = {}
    for att in attendances:
        att_map[(att.student_id, att.session_id)] = att

    rows = []
    matrix = {}
    for student in students:
        student_key = str(student.pk)
        matrix[student_key] = {}
        present_count = 0
        for session in sessions:
            att = att_map.get((student.pk, session.pk))
            present = att is not None
            if present:
                present_count += 1
            local_time = (
                timezone.localtime(att.validation_time).strftime("%Y-%m-%d %H:%M:%S")
                if att
                else None
            )
            matrix[student_key][str(session.pk)] = {
                "present": present,
                "validation_time": local_time,
            }
            rows.append(
                {
                    "student_id": student.pk,
                    "session_id": session.pk,
                    "first_name": student.first_name,
                    "last_name": student.last_name,
                    "code_massar": student.code_massar,
                    "classe": classe.name,
                    "subject": session.subject,
                    "date": str(session.date),
                    "start_time": str(session.start_time)[:5],
                    "end_time": str(session.end_time)[:5],
                    "present": present,
                    "status_label": "Présent" if present else "Absent",
                    "validation_time": local_time,
                }
            )

        matrix[student_key]["_summary"] = {
            "present_count": present_count,
            "absent_count": len(sessions) - present_count,
            "rate": round((present_count / len(sessions)) * 100, 1) if sessions else 0,
        }

    return {
        "classe": {"id": classe.pk, "name": classe.name, "code": classe.code},
        "sessions": [
            {
                "id": s.pk,
                "subject": s.subject,
                "date": str(s.date),
                "start_time": str(s.start_time)[:5],
                "end_time": str(s.end_time)[:5],
            }
            for s in sessions
        ],
        "students": [
            {
                "id": s.pk,
                "first_name": s.first_name,
                "last_name": s.last_name,
                "code_massar": s.code_massar,
            }
            for s in students
        ],
        "matrix": matrix,
        "rows": rows,
    }


def _style_excel_header(ws):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    return ws


def _autosize_columns(ws):
    for column in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 40)


def _build_custom_export_workbook(data):
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    sessions = data["sessions"]
    students = data["students"]
    matrix = data["matrix"]

    ws_summary = wb.active
    ws_summary.title = "Récapitulatif"
    summary_headers = [
        "Prénom",
        "Nom",
        "Code Massar",
        "Classe",
        *[
            f"{s['subject']} ({s['date']})"
            for s in sessions
        ],
        "Présences",
        "Absences",
        "Taux (%)",
    ]
    ws_summary.append(summary_headers)
    _style_excel_header(ws_summary)

    present_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    absent_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    for student in students:
        sid = str(student["id"])
        summary = matrix[sid]["_summary"]
        row = [
            student["first_name"],
            student["last_name"],
            student["code_massar"],
            data["classe"]["name"],
        ]
        for session in sessions:
            cell_data = matrix[sid][str(session["id"])]
            row.append("Présent" if cell_data["present"] else "Absent")
        row.extend(
            [
                summary["present_count"],
                summary["absent_count"],
                summary["rate"],
            ]
        )
        ws_summary.append(row)
        row_idx = ws_summary.max_row
        for col_idx, session in enumerate(sessions, start=5):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.fill = present_fill if cell.value == "Présent" else absent_fill

    _autosize_columns(ws_summary)

    ws_detail = wb.create_sheet("Détail")
    detail_headers = [
        "Prénom",
        "Nom",
        "Code Massar",
        "Classe",
        "Matière",
        "Date séance",
        "Heure début",
        "Heure fin",
        "Statut",
        "Heure validation",
    ]
    ws_detail.append(detail_headers)
    _style_excel_header(ws_detail)

    for row in data["rows"]:
        ws_detail.append(
            [
                row["first_name"],
                row["last_name"],
                row["code_massar"],
                row["classe"],
                row["subject"],
                row["date"],
                row["start_time"],
                row["end_time"],
                row["status_label"],
                row["validation_time"] or "",
            ]
        )
    _autosize_columns(ws_detail)

    ws_info = wb.create_sheet("Informations")
    ws_info.append(["Classe", data["classe"]["name"]])
    ws_info.append(["Code classe", data["classe"]["code"]])
    ws_info.append(["Nombre d'étudiants", len(students)])
    ws_info.append(["Nombre de séances", len(sessions)])
    ws_info.append(["Export généré le", timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")])
    for cell in ws_info["A"]:
        cell.font = Font(bold=True)
    _autosize_columns(ws_info)

    return wb


class CustomExportPreviewView(APIView):
    """Preview personalized attendance report for a class."""

    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request):
        classe_id = request.query_params.get("classe_id")
        if not classe_id:
            return Response(
                {"error": "classe_id requis."}, status=status.HTTP_400_BAD_REQUEST
            )

        session_ids = _parse_id_list(request.query_params.get("session_ids", ""))
        student_ids = _parse_id_list(request.query_params.get("student_ids", ""))

        try:
            data = _build_custom_export_data(
                request.user,
                int(classe_id),
                session_ids=session_ids or None,
                student_ids=student_ids or None,
            )
        except PermissionError as e:
            return Response({"error": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(data)


class CustomExportView(APIView):
    """Export personalized attendance report to Excel."""

    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def get(self, request):
        classe_id = request.query_params.get("classe_id")
        if not classe_id:
            return Response(
                {"error": "classe_id requis."}, status=status.HTTP_400_BAD_REQUEST
            )

        session_ids = _parse_id_list(request.query_params.get("session_ids", ""))
        student_ids = _parse_id_list(request.query_params.get("student_ids", ""))

        try:
            data = _build_custom_export_data(
                request.user,
                int(classe_id),
                session_ids=session_ids or None,
                student_ids=student_ids or None,
            )
        except PermissionError as e:
            return Response({"error": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        wb = _build_custom_export_workbook(data)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_name = "".join(
            c if c.isalnum() else "_" for c in data["classe"]["name"]
        )[:30]
        filename = f"export_personnalise_{safe_name}_{timezone.localdate()}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        logger.info(
            "Custom export for classe %s by %s (%s students, %s sessions)",
            classe_id,
            request.user.email,
            len(data["students"]),
            len(data["sessions"]),
        )
        return response
