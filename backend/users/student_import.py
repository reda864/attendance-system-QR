import re
from io import BytesIO

import openpyxl
from django.db import transaction

from .models import Classe, Student

HEADER_MAP = {
    "first_name": {
        "prenom",
        "prénom",
        "pre nom",
        "first name",
        "firstname",
        "first_name",
    },
    "last_name": {
        "nom",
        "last name",
        "lastname",
        "last_name",
        "nom de famille",
        "family name",
    },
    "code_massar": {
        "massar",
        "code massar",
        "code_massar",
        "codemassar",
        "code-massar",
        "cne",
    },
    "classe": {
        "classe",
        "class",
        "classe_name",
        "groupe",
        "group",
        "classe name",
    },
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _map_headers(row_values) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row_values):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        for field, aliases in HEADER_MAP.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _cell_value(row, index: int | None):
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _resolve_classe(raw_value: str, classes_by_name: dict[str, Classe]) -> Classe | None:
    if not raw_value:
        return None
    key = raw_value.strip().lower()
    if key.isdigit():
        try:
            return Classe.objects.get(pk=int(key))
        except Classe.DoesNotExist:
            pass
    return classes_by_name.get(key)


def parse_student_rows(file_obj) -> tuple[list[dict], list[dict]]:
    """Return (rows, parse_errors) from an Excel file."""
    try:
        workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        return [], [{"row": 0, "message": f"Fichier Excel invalide : {exc}"}]

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return [], [{"row": 1, "message": "Le fichier est vide."}]

    column_map = _map_headers(header_row)
    required = {"first_name", "last_name", "code_massar", "classe"}
    missing = required - set(column_map)
    if missing:
        labels = {
            "first_name": "Prénom",
            "last_name": "Nom",
            "code_massar": "Massar",
            "classe": "Classe",
        }
        workbook.close()
        return [], [
            {
                "row": 1,
                "message": (
                    "Colonnes manquantes : "
                    + ", ".join(labels[field] for field in sorted(missing))
                    + ". Attendu : Nom, Prénom, Massar, Classe."
                ),
            }
        ]

    parsed_rows: list[dict] = []
    parse_errors: list[dict] = []
    for row_number, row in enumerate(rows_iter, start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        parsed_rows.append(
            {
                "row": row_number,
                "first_name": _cell_value(row, column_map["first_name"]),
                "last_name": _cell_value(row, column_map["last_name"]),
                "code_massar": _cell_value(row, column_map["code_massar"]),
                "classe": _cell_value(row, column_map["classe"]),
            }
        )

    workbook.close()
    return parsed_rows, parse_errors


def import_students_from_excel(file_obj) -> dict:
    rows, parse_errors = parse_student_rows(file_obj)
    if parse_errors:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": parse_errors}

    classes_by_name = {
        c.name.strip().lower(): c for c in Classe.objects.only("id", "name")
    }
    existing_massar = set(
        Student.objects.values_list("code_massar", flat=True)
    )

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    with transaction.atomic():
        for row in rows:
            row_number = row["row"]
            first_name = row["first_name"]
            last_name = row["last_name"]
            code_massar = row["code_massar"]
            classe_raw = row["classe"]

            if not first_name or not last_name or not code_massar or not classe_raw:
                errors.append(
                    {
                        "row": row_number,
                        "message": "Nom, prénom, Massar et classe sont obligatoires.",
                    }
                )
                continue

            classe = _resolve_classe(classe_raw, classes_by_name)
            if not classe:
                errors.append(
                    {
                        "row": row_number,
                        "message": f"Classe introuvable : « {classe_raw} ».",
                    }
                )
                continue

            if code_massar in existing_massar:
                student = Student.objects.filter(code_massar=code_massar).first()
                if student:
                    student.first_name = first_name
                    student.last_name = last_name
                    student.classe = classe
                    student.save(update_fields=["first_name", "last_name", "classe"])
                    updated += 1
                else:
                    skipped += 1
                continue

            Student.objects.create(
                first_name=first_name,
                last_name=last_name,
                code_massar=code_massar,
                email="",
                phone="",
                classe=classe,
            )
            existing_massar.add(code_massar)
            created += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def build_import_template() -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etudiants"
    ws.append(["Nom", "Prénom", "Massar", "Classe"])
    ws.append(["Benali", "Ayman", "M123456789", "Genie Informatique - S5 - 2026-2027"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
